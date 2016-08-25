# coding=utf-8
from openpyxl import Workbook
from io import BytesIO


def create_sheet(title, header_list, datas):
    '''
    parameter:
        title: str
        header_list: []
        datas: [[], []]
    '''
    count = 0
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet()
    sheet.title = title
    sheet.append(header_list)
    for data in datas:
        sheet.append(data)
        count += 1
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
